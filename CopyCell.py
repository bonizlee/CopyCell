# CopyCell.py
# BonizLee
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os.path
import os
import sys
import re
import shutil

#初始化
def init(argv=None):
    alen = len(argv)
    global PATH
    PATH = os.path.dirname(os.path.realpath(__file__))+os.path.sep
    global EXECLFILE
    #EXECLFILE='省厅认定清单20171227.xlsx'
    global SOURCETEMPLATE  #数据源文件
    #SOURCETEMPLATE='ndhs_in_2017排放分类统计_'
    global TARGETTEMPLATE  #模板文件
    global SOURCESHEET  #数据源文件
    global TARGETSHEET  #模板文件
    #TARGETTEMPLATE='排放统计'
    global WITHSHEETNAME
    WITHSHEETNAME = False

    if alen==3:
        SOURCETEMPLATE=argv[1]
        TARGETTEMPLATE=argv[2]
    elif alen==5:
        WITHSHEETNAME=True
        SOURCETEMPLATE=argv[1]
        SOURCESHEET=argv[2]
        TARGETTEMPLATE=argv[3]
        TARGETSHEET=argv[4]
    else:
        print('命令格式错误。正例(不需要扩展名，只支持xlsx，表名大小写敏感)：命令  源文件名模板  [源电子表名]  模板文件名  [模板电子表名]')
        return False
    return True

    '''
    global FZJG_DIC
    FZJG_DIC={ '粤A':'广州','粤B':'深圳','粤C':'珠海','粤D':'汕头','粤E':'佛山','粤F':'韶关','粤G':'湛江','粤H':'肇庆','粤J':'江门','粤K':'茂名','粤L':'惠州','粤M':'梅州','粤N':'汕尾','粤P':'河源','粤Q':'阳江','粤R':'清远','粤S':'东莞','粤T':'中山','粤U':'潮州','粤V':'揭阳','粤W':'云浮'}
    '''

def get_source_files():
    srclist = {}
    pstr='('+SOURCETEMPLATE+')(.*)'
    for root, dirs, files in os.walk(PATH):
        for f in files:
            if os.path.splitext(f)[1] == '.xlsx':
                fname=os.path.splitext(f)[0]
                pattern=re.compile(pstr)
                m=pattern.match(fname)
                if not m is None:
                    pos=len(SOURCETEMPLATE)
                    k=fname[pos:]                    
                    srclist.setdefault(k,f)   
    return srclist

def copy_excel(source=None,target=None):
    srcWB=load_workbook(PATH+source)
    tarWB=load_workbook(PATH+target)
    if not WITHSHEETNAME:
        srcWS=srcWB.get_active_sheet()
        tarWS=tarWB.get_active_sheet()
    else:
        srcWS=srcWB.get_sheet_by_name(SOURCESHEET)
        tarWS=tarWB.get_sheet_by_name(TARGETSHEET)

    #rows_len = srcWS.max_row
    #两个for循环遍历整个excel的单元格内容
    for i,row in enumerate(srcWS.iter_rows()):
        for j,cell in enumerate(row):
            tarWS.cell(row=i+1, column=j+1, value=cell.value)
    tarWB.save(target)

def exec_copy(srcdict=None):
    for k,v in srcdict.items():
        targetname=TARGETTEMPLATE+'_'+k+'.xlsx'
        shutil.copyfile(TARGETTEMPLATE+'.xlsx',targetname)
        copy_excel(v,targetname)

if __name__ == "__main__":
    if init(sys.argv):
        try:
            srcdict=get_source_files()
            exec_copy(srcdict)
        except FileNotFoundError:
            print('文件名或路径错误')
        except KeyError:
            print('Sheet表名不正确，注意大小写敏感')
        print('完成')